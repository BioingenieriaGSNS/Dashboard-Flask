from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
import psycopg2
import psycopg2.extras
import os
from dotenv import load_dotenv
from datetime import datetime, date

# Importar el sistema de autenticación

from auth import (
    login_manager, User, authenticate_user, create_user, 
    permission_required, role_required, get_all_users,
    toggle_user_status, update_user_role, update_user_password,
    update_last_login, update_own_password, get_user_profile  # 👈 Nuevos
)
load_dotenv()

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'cambiar-esta-clave-secreta-en-produccion')

# Configurar Flask-Login
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = '⚠️ Por favor inicia sesión para acceder'
login_manager.login_message_category = 'error'

DATABASE_URL = os.getenv('DATABASE_URL')

def get_db_connection():
    """Conexión a la base de datos"""
    try:
        db_url_clean = DATABASE_URL.replace('&channel_binding=require', '')
        conn = psycopg2.connect(db_url_clean, cursor_factory=psycopg2.extras.RealDictCursor)
        return conn
    except Exception as e:
        print(f"Error de conexión: {e}")
        return None

def registrar_auditoria(conn, equipo_id, usuario_id, usuario_nombre, campo, valor_anterior, valor_nuevo, accion='UPDATE'):
    """Registra un cambio en la tabla de auditoría"""
    try:
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO equipos_auditoria 
            (equipo_id, usuario_id, usuario_nombre, campo_modificado, valor_anterior, valor_nuevo, accion)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (equipo_id, usuario_id, usuario_nombre, campo, str(valor_anterior) if valor_anterior else '', str(valor_nuevo) if valor_nuevo else '', accion))
        cursor.close()
    except Exception as e:
        print(f"Error al registrar auditoría: {e}")

# ============================================
# RUTAS DE AUTENTICACIÓN
# ============================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = authenticate_user(username, password)
        
        if user:
            login_user(user)
            update_last_login(user.id)
            flash(f'✅ Bienvenido {user.username}!', 'success')
            
            # Redirigir a la página que intentaban acceder o al dashboard
            next_page = request.args.get('next')
            return redirect(next_page if next_page else url_for('index'))
        else:
            flash('❌ Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Cerrar sesión"""
    logout_user()
    flash('👋 Has cerrado sesión exitosamente', 'success')
    return redirect(url_for('login'))

# ============================================
# RUTAS PRINCIPALES (REQUIEREN AUTENTICACIÓN)
# ============================================

@app.route('/')
@login_required
def index():
    """Página principal - Dashboard"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión a la base de datos", 500
    
    cursor = conn.cursor()
    
    # Métricas totales
    cursor.execute("SELECT COUNT(*) as total FROM solicitudes")
    total_solicitudes = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as total FROM equipos WHERE eliminado = FALSE")
    total_equipos = cursor.fetchone()['total']
    
    # Métricas de estados específicos para las tarjetas superiores
    cursor.execute("SELECT COUNT(*) as total FROM equipos WHERE estado = 'Pendiente' AND eliminado = FALSE")
    pendientes = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as total FROM equipos WHERE estado = 'Finalizado' AND eliminado = FALSE")
    finalizados = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as total FROM equipos WHERE estado = 'En curso' AND eliminado = FALSE")
    en_curso = cursor.fetchone()['total']
    
    cursor.execute("SELECT COUNT(*) as total FROM equipos WHERE estado = 'A presupuestar' AND eliminado = FALSE")
    a_presupuestar = cursor.fetchone()['total']
    
    # Estados de solicitudes (todos los estados)
    cursor.execute("""
        SELECT estado, COUNT(*) as cantidad 
        FROM equipos 
        WHERE eliminado = FALSE
        GROUP BY estado
        ORDER BY 
            CASE estado
                WHEN 'Pendiente' THEN 1
                WHEN 'Aprobación pendiente' THEN 2
                WHEN 'A presupuestar' THEN 3
                WHEN 'Baja técnica' THEN 4
                WHEN 'En curso' THEN 5
                WHEN 'Finalizado' THEN 6
                WHEN 'Listo para entregar' THEN 7
                WHEN 'Repuestos' THEN 8
                WHEN 'Tercerizado' THEN 9
                ELSE 11
            END
    """)
    estados = cursor.fetchall()
    
    # Categorías de equipos basadas en la columna categoria de solicitudes
    cursor.execute("""
        SELECT 
            CASE 
                WHEN s.categoria LIKE '%R%' THEN 'Reparación'
                WHEN s.categoria LIKE '%G%' THEN 'Garantía'
                WHEN s.categoria LIKE '%BA%' THEN 'Baja de Alquiler'
                WHEN s.categoria LIKE '%CA%' THEN 'Cambio de Alquiler'
                WHEN s.categoria LIKE '%FC%' THEN 'Cambio por Falla Crítica'
                ELSE 'Otra'
            END as categoria_nombre,
            COUNT(*) as cantidad
        FROM equipos e
        LEFT JOIN solicitudes s ON e.solicitud_id = s.id
        WHERE e.eliminado = FALSE
        GROUP BY categoria_nombre
        ORDER BY cantidad DESC
    """)
    categorias = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('dashboard.html', 
                         total_solicitudes=total_solicitudes,
                         total_equipos=total_equipos,
                         pendientes=pendientes,
                         finalizados=finalizados,
                         en_curso=en_curso,
                         a_presupuestar=a_presupuestar,
                         estados=estados,
                         categorias=categorias)

@app.route('/solicitudes')
@login_required
def solicitudes():
    """Página de solicitudes"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión a la base de datos", 500
    
    cursor = conn.cursor()
    cursor.execute("""
    SELECT 
        s.id,
        s.fecha_solicitud,
        s.estado,
        s.categoria,
        s.pdf_url,
        s.email_solicitante,
        s.quien_completa,
        s.nivel_urgencia,
        s.motivo_solicitud,
        s.comercial_syemed,
        CASE WHEN s.categoria LIKE '%G%' THEN 'Sí' ELSE 'No' END as garantia,
        -- Colaborador Syemed
        s.area_solicitante,
        s.solicitante,
        s.logistica_cargo,
        s.comentarios_caso,
        s.equipo_corresponde_a,
        -- Distribuidor / Institución
        s.nombre_fantasia,
        s.razon_social,
        s.cuit,
        s.contacto_nombre,
        s.contacto_telefono,
        s.contacto_tecnico,
        s.equipo_propiedad,
        -- Paciente Particular
        s.nombre_apellido_paciente,
        s.telefono_paciente,
        s.equipo_origen,
        -- OSTs vinculadas
        (SELECT STRING_AGG(DISTINCT e.ost::TEXT, ', ' ORDER BY e.ost::TEXT)
         FROM equipos e 
         WHERE e.solicitud_id = s.id AND e.eliminado = FALSE) as osts_vinculadas 
    FROM solicitudes s
    ORDER BY s.fecha_solicitud DESC
""")
    solicitudes = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return render_template('solicitudes.html', solicitudes=solicitudes)


@app.route('/equipos')
@login_required
def equipos():
    """Página de equipos (solo muestra equipos NO eliminados)"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión a la base de datos", 500
    
    cursor = conn.cursor()
    
    # Obtener equipos NO eliminados
    cursor.execute("""
        SELECT e.id, e.cliente, e.ost, e.estado, e.fecha_ingreso, e.remito,
            e.tipo_equipo, e.marca, e.modelo, e.numero_serie, e.accesorios,
            s.categoria,
            COALESCE(s.comercial_syemed, s.solicitante) as comercial_cargo,
            e.observacion_ingreso, e.prioridad, e.fecha_envio, e.proveedor,
            e.detalles_reparacion, e.horas_trabajo, e.reingreso, 
            e.informe AS informe_tecnico,
            e.costo AS costo_reparacion, 
            e.precio AS precio_cliente, 
            e.ov AS numero_ov, 
            e.estado_ov, e.fecha_entrega, e.remito_entrega,
            e.solicitud_id,
            s.nivel_urgencia
        FROM equipos e
        LEFT JOIN solicitudes s ON e.solicitud_id = s.id
        WHERE e.eliminado = FALSE
        ORDER BY e.fecha_ingreso DESC
    """)
    equipos = cursor.fetchall()
    
    # Obtener archivos para las fotos (hacer JOIN con equipos NO eliminados)
    cursor.execute("""
        SELECT e.numero_serie, a.categoria, a.url_cloudinary
        FROM archivos_adjuntos a
        INNER JOIN equipos e ON a.equipo_id = e.id
        WHERE e.numero_serie IS NOT NULL 
        AND e.eliminado = FALSE  -- 👈 Solo archivos de equipos activos
    """)
    archivos = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('equipos.html', equipos=equipos, archivos=archivos)

@app.route('/archivos')
@login_required
def archivos():
    """Página de archivos adjuntos"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión a la base de datos", 500
    
    cursor = conn.cursor()
    cursor.execute("""
        SELECT a.id, a.solicitud_id, a.equipo_id, a.nombre_archivo,
               a.url_cloudinary, a.tipo_archivo, a.categoria,
               a.tamano_bytes, a.fecha_subida,
               e.ost, e.numero_serie
        FROM archivos_adjuntos a
        LEFT JOIN equipos e ON a.equipo_id = e.id
        ORDER BY a.fecha_subida DESC
    """)
    archivos = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return render_template('archivos.html', archivos=archivos)

# ============================================
# GESTIÓN DE USUARIOS (SOLO ADMIN)
# ============================================

@app.route('/usuarios')
@permission_required('manage_users')
def usuarios():
    """Página de gestión de usuarios (solo admin)"""
    users = get_all_users()
    return render_template('usuarios.html', users=users)

@app.route('/perfil')
@login_required
def perfil():
    """Página de perfil del usuario"""
    from auth import get_user_profile
    user_data = get_user_profile(current_user.id)
    return render_template('perfil.html', user_data=user_data)

@app.route('/api/perfil/cambiar-password', methods=['POST'])
@login_required
def cambiar_password_perfil():
   
    if not current_user.has_permission('edit'):
        return jsonify({'success': False, 'error': 'No tienes permiso para cambiar tu contraseña'}), 403    
def cambiar_mi_password():
    """API para que un usuario cambie su propia contraseña"""
    from auth import update_own_password
    try:
        data = request.json
        update_own_password(
            current_user.id,
            data['current_password'],
            data['new_password']
        )
        return jsonify({'success': True, 'message': 'Contraseña actualizada correctamente'})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# ============================================
# AUDITORÍA (SOLO ADMIN)
# ============================================


@app.route('/auditoria')
@permission_required('view_audit')
def auditoria():
    """Página de auditoría de cambios (solo admin)"""
    conn = get_db_connection()
    if not conn:
        return "Error de conexión a la base de datos", 500
    
    cursor = conn.cursor()
    
    # Obtener el equipo_id si se filtra
    equipo_id = request.args.get('equipo_id', type=int)
    
    if equipo_id:
        cursor.execute("""
            SELECT a.*, e.ost, e.cliente, e.tipo_equipo, e.eliminado
            FROM equipos_auditoria a
            LEFT JOIN equipos e ON a.equipo_id = e.id
            WHERE a.equipo_id = %s
            ORDER BY a.fecha_cambio DESC
        """, (equipo_id,))
    else:
        # 👇 CLAVE: No filtramos por eliminado, mostramos TODO
        cursor.execute("""
            SELECT a.*, e.ost, e.cliente, e.tipo_equipo, e.eliminado
            FROM equipos_auditoria a
            LEFT JOIN equipos e ON a.equipo_id = e.id
            ORDER BY a.fecha_cambio DESC
            LIMIT 500
        """)
    
    cambios = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return render_template('auditoria.html', cambios=cambios, equipo_id=equipo_id)
# ============================================
# API ENDPOINTS
# ============================================

@app.route('/api/solicitud/<int:id>', methods=['PUT'])
@permission_required('edit')
def update_solicitud(id):
    """API para actualizar solicitud"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión'}), 500
    
    data = request.json
    cursor = conn.cursor()
    
    try:
        campos = []
        valores = []
        
        # Lista de campos actualizables (ahora incluye todos los campos)
        campos_permitidos = [
            # Datos de Ingreso
            'categoria', 'email_solicitante', 'quien_completa', 
            'nivel_urgencia', 'motivo_solicitud', 'comercial_cargo', 'estado',
            # Colaborador Syemed
            'area_solicitante', 'solicitante', 'logistica_cargo',
            'comentarios_caso', 'equipo_corresponde_a',
            # Distribuidor / Institución
            'nombre_fantasia', 'razon_social', 'cuit', 
            'nombre_contacto', 'contacto_telefono', 'contacto_tecnico',
            'equipo_propiedad',
            # Paciente Particular
            'nombre_apellido_paciente', 'telefono_paciente', 'equipo_origen'
        ]
        
        for campo in campos_permitidos:
            if campo in data:
                campos.append(f'{campo} = %s')
                valores.append(data[campo])
        
        if not campos:
            return jsonify({'error': 'No hay campos para actualizar'}), 400
        
        valores.append(id)
        query = f"UPDATE solicitudes SET {', '.join(campos)} WHERE id = %s"
        
        cursor.execute(query, valores)
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/api/proximo-ost', methods=['GET'])
@login_required
def obtener_proximo_ost():
    """API para obtener el próximo número OST disponible"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión'}), 500
    
    cursor = conn.cursor()
    
    try:
        # Obtener el último OST
        cursor.execute("SELECT MAX(ost) as max_ost FROM equipos")
        result = cursor.fetchone()
        max_ost = result['max_ost'] if result and result['max_ost'] else 0
        proximo_ost = max_ost + 1
        
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'proximo_ost': proximo_ost
        })
    except Exception as e:
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/equipos', methods=['POST'])
@app.route('/api/equipo/crear', methods=['POST'])
@permission_required('edit')
def crear_equipo():
    """API para crear nuevo equipo"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión'}), 500
    
    data = request.json
    cursor = conn.cursor()
    
    try:
        fecha_ingreso = data.get('fecha_ingreso')
        if isinstance(fecha_ingreso, str):
            try:
                fecha_ingreso = datetime.strptime(fecha_ingreso, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'error': 'Formato de fecha inválido'}), 400
        
        def empty_to_none(value):
            return None if value == '' or value is None else value
        
        cursor.execute("""
            INSERT INTO equipos (
                cliente, tipo_equipo, marca, modelo, numero_serie,
                fecha_ingreso, remito, accesorios, prioridad, 
                observacion_ingreso, estado
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id, ost
        """, (
            data.get('cliente'),
            data.get('tipo_equipo'),
            empty_to_none(data.get('marca')),
            empty_to_none(data.get('modelo')),
            empty_to_none(data.get('numero_serie')),
            fecha_ingreso,
            empty_to_none(data.get('remito')),
            empty_to_none(data.get('accesorios')),
            data.get('prioridad', 'Media'),
            empty_to_none(data.get('observacion_ingreso')),
            'Pendiente'
        ))
        result = cursor.fetchone()
        
        # Registrar en auditoría
        registrar_auditoria(
            conn, 
            result['id'], 
            current_user.id, 
            current_user.username,
            'CREACIÓN',
            '',
            f"OST: {result['ost']}, Cliente: {data.get('cliente')}",
            'INSERT'
        )
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'id': result['id'],
            'ost': result['ost'],
            'message': 'Equipo creado exitosamente'
        }), 201
    
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/equipo/<int:id>', methods=['PUT'])
@permission_required('edit')
def update_equipo(id):
    """API para actualizar equipo (requiere permiso de edición)"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión'}), 500
    
    data = request.json
    cursor = conn.cursor()
    
    try:
        # Primero obtener los valores actuales
        cursor.execute("SELECT * FROM equipos WHERE id = %s", (id,))
        equipo_actual = cursor.fetchone()
        
        if not equipo_actual:
            return jsonify({'error': 'Equipo no encontrado'}), 404
        
        # Construir la consulta dinámicamente solo con los campos que vienen
        campos = []
        valores = []
        
        # Mapeo de campos JSON a columnas DB
        campo_map = {
            'cliente': 'cliente',
            'tipo_equipo': 'tipo_equipo',
            'marca': 'marca',
            'modelo': 'modelo',
            'numero_serie': 'numero_serie',
            'accesorios': 'accesorios',
            'prioridad': 'prioridad',
            'remito': 'remito',
            'observacion_ingreso': 'observacion_ingreso',
            'detalle_reparacion': 'detalles_reparacion',
            'horas_trabajo': 'horas_trabajo',
            'reingreso': 'reingreso',
            'informe_tecnico': 'informe',
            'costo_reparacion': 'costo',
            'precio_cliente': 'precio',
            'numero_ov': 'ov',
            'estado_ov': 'estado_ov',
            'fecha_ingreso': 'fecha_ingreso',
            'fecha_envio_proveedor': 'fecha_envio',
            'fecha_entrega': 'fecha_entrega',
            'remito_entrega': 'remito_entrega',
            'estado': 'estado',
            'proveedor': 'proveedor'
        }
        
        for campo_json, campo_db in campo_map.items():
            if campo_json in data:
                valor_anterior = equipo_actual.get(campo_db)
                valor_nuevo = data.get(campo_json)
                
                # Solo actualizar si el valor cambió
                if str(valor_anterior) != str(valor_nuevo):
                    campos.append(f'{campo_db} = %s')
                    valores.append(valor_nuevo)
                    
                    # Registrar en auditoría
                    registrar_auditoria(
                        conn,
                        id,
                        current_user.id,
                        current_user.username,
                        campo_db,
                        valor_anterior,
                        valor_nuevo
                    )
        
        if not campos:
            return jsonify({'success': True, 'message': 'No hay cambios para guardar'})
        
        # Agregar el ID al final
        valores.append(id)
        
        query = f"UPDATE equipos SET {', '.join(campos)} WHERE id = %s"
        
        cursor.execute(query, valores)
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error al actualizar: {e}")
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'error': str(e)}), 500

# REEMPLAZA los endpoints de DELETE y RESTAURAR en app.py con estos:

@app.route('/api/equipo/<int:id>', methods=['DELETE'])
@permission_required('delete')
def delete_equipo(id):
    """API para eliminar equipo (Soft Delete - requiere permiso de eliminación)"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión'}), 500
    
    cursor = conn.cursor()
    
    try:
        # Obtener datos del equipo antes de "eliminar"
        cursor.execute("SELECT * FROM equipos WHERE id = %s AND eliminado = FALSE", (id,))
        equipo = cursor.fetchone()
        
        if not equipo:
            return jsonify({'success': False, 'error': 'Equipo no encontrado o ya está eliminado'}), 404
        
        # Marcar como eliminado (Soft Delete)
        cursor.execute("""
            UPDATE equipos 
            SET eliminado = TRUE, fecha_eliminacion = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (id,))
        
        # Registrar la eliminación en auditoría
        registrar_auditoria(
            conn, 
            id, 
            current_user.id, 
            current_user.username,
            'ELIMINACIÓN',
            f"OST: {equipo['ost']}, Cliente: {equipo['cliente']}, Tipo: {equipo['tipo_equipo']}",
            'EQUIPO ELIMINADO (SOFT DELETE)',
            'DELETE'
        )
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Equipo eliminado correctamente'})
    except Exception as e:
        print(f"Error al eliminar: {e}")
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/equipo/<int:id>/restaurar', methods=['POST'])
@permission_required('delete')
def restaurar_equipo(id):
    """API para restaurar un equipo eliminado (requiere permiso de eliminación)"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Error de conexión'}), 500
    
    cursor = conn.cursor()
    
    try:
        # Buscar el equipo eliminado
        cursor.execute("""
            SELECT * FROM equipos 
            WHERE id = %s AND eliminado = TRUE
        """, (id,))
        
        equipo = cursor.fetchone()
        
        if not equipo:
            return jsonify({
                'success': False, 
                'error': 'No se encontró el equipo eliminado'
            }), 404
        
        # Restaurar el equipo (marcar eliminado = FALSE)
        cursor.execute("""
            UPDATE equipos 
            SET eliminado = FALSE, fecha_eliminacion = NULL
            WHERE id = %s
        """, (id,))
        
        # Registrar la restauración en auditoría
        registrar_auditoria(
            conn,
            id,
            current_user.id,
            current_user.username,
            'RESTAURACIÓN',
            f'Equipo eliminado - OST: {equipo["ost"]}, Cliente: {equipo["cliente"]}',
            f'Equipo restaurado con OST: {equipo["ost"]}',
            'INSERT'
        )
        
        conn.commit()
        cursor.close()
        conn.close()
        
        return jsonify({
            'success': True,
            'id': equipo['id'],
            'ost': equipo['ost'],
            'message': 'Equipo restaurado exitosamente'
        })
    
    except Exception as e:
        print(f"Error al restaurar: {e}")
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500
    
# ============================================
# API ENDPOINTS PARA GESTIÓN DE USUARIOS (SOLO ADMIN)
# ============================================

@app.route('/api/users/create', methods=['POST'])
@permission_required('manage_users')
def api_create_user():
    """API para crear usuario (solo admin)"""
    try:
        data = request.json
        user_id = create_user(
            username=data['username'],
            email=data['email'],
            password=data['password'],
            role=data['role']
        )
        return jsonify({'success': True, 'user_id': user_id})
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users/update-role', methods=['POST'])
@permission_required('manage_users')
def api_update_role():
    """API para actualizar rol de usuario (solo admin)"""
    try:
        data = request.json
        update_user_role(data['user_id'], data['role'])
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users/update-password', methods=['POST'])
@permission_required('manage_users')
def api_update_password():
    """API para actualizar contraseña de usuario (solo admin)"""
    try:
        data = request.json
        update_user_password(data['user_id'], data['new_password'])
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/users/toggle-status', methods=['POST'])
@permission_required('manage_users')
def api_toggle_status():
    """API para activar/desactivar usuario (solo admin)"""
    try:
        data = request.json
        toggle_user_status(data['user_id'])
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    
@app.route('/equipos-priorizados')
@login_required
@permission_required('view')
def equipos_priorizados():
    """Muestra equipos priorizados - disponible para todos menos viewer"""
    if current_user.role == 'viewer':
        flash('⛔ No tienes permiso para ver los equipos priorizados', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    cur = conn.cursor()
    
    try:
        # Obtener equipos priorizados
        cur.execute("SELECT * FROM equipos_priorizados")
        
        equipos = cur.fetchall()
        
        # Agrupar por nivel de prioridad
        equipos_agrupados = {
            'Critica': [],
            'Alta': [],
            'Media': [],
            'Baja': []
        }
        
        for equipo in equipos:
            nivel = equipo['nivel_prioridad']
            equipos_agrupados[nivel].append(equipo)
        
        # Estadísticas
        stats = {
            'total': len(equipos),
            'critica': len(equipos_agrupados['Critica']),
            'alta': len(equipos_agrupados['Alta']),
            'media': len(equipos_agrupados['Media']),
            'baja': len(equipos_agrupados['Baja'])
        }
        
        return render_template('equipos_priorizados.html', 
                             equipos_agrupados=equipos_agrupados,
                             stats=stats)
    
    except Exception as e:
        flash(f'Error al cargar equipos priorizados: {str(e)}', 'error')
        return redirect(url_for('index'))
    finally:
        cur.close()
        conn.close()

@app.route('/informes-mensuales')
@login_required
def informes_mensuales():
    """Página de informes mensuales"""
    return render_template('informes_mensuales.html')

@app.route('/api/equipo/detalle/<ost>')
@login_required
def api_equipo_detalle(ost):
    """API para obtener detalles completos de un equipo por OST"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Error de conexión'}), 500
    
    cursor = conn.cursor()
    
    try:
        cursor.execute("""
            SELECT 
                e.id, e.cliente, e.ost, e.estado, e.fecha_ingreso, e.remito,
                e.tipo_equipo, e.marca, e.modelo, e.numero_serie, e.accesorios,
                s.categoria,
                COALESCE(s.comercial_syemed, s.solicitante) as comercial_cargo,
                e.observacion_ingreso, e.prioridad, e.fecha_envio, e.proveedor,
                e.detalles_reparacion, e.horas_trabajo, e.reingreso, 
                e.informe AS informe_tecnico,
                e.costo AS costo_reparacion, 
                e.precio AS precio_cliente, 
                e.ov AS numero_ov, 
                e.estado_ov, e.fecha_entrega, e.remito_entrega,
                e.solicitud_id,
                s.nivel_urgencia
            FROM equipos e
            LEFT JOIN solicitudes s ON e.solicitud_id = s.id
            WHERE e.ost = %s AND e.eliminado = FALSE
            LIMIT 1
        """, (ost,))
        
        equipo = cursor.fetchone()
        
        if not equipo:
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': 'Equipo no encontrado'}), 404
        
        # Convertir RealDictRow a dict normal y formatear fechas
        equipo_dict = dict(equipo)
        
        # Formatear fechas
        if equipo_dict.get('fecha_ingreso'):
            equipo_dict['fecha_ingreso'] = equipo_dict['fecha_ingreso'].strftime('%d/%m/%Y')
        if equipo_dict.get('fecha_envio'):
            equipo_dict['fecha_envio'] = equipo_dict['fecha_envio'].strftime('%d/%m/%Y')
        if equipo_dict.get('fecha_entrega'):
            equipo_dict['fecha_entrega'] = equipo_dict['fecha_entrega'].strftime('%d/%m/%Y')
        
        cursor.close()
        conn.close()
        
        return jsonify({'success': True, 'equipo': equipo_dict})
    
    except Exception as e:
        print(f"Error al obtener equipo: {e}")
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/informe-mensual')
@login_required
def api_informe_mensual():
    """API para generar informe mensual de equipos"""
    anio = request.args.get('anio', type=int)
    mes = request.args.get('mes', type=int)
    categoria = request.args.get('categoria', '')
    
    if not anio:
        return jsonify({'success': False, 'error': 'Año requerido'}), 400
    
    conn = get_db_connection()
    if not conn:
        return jsonify({'success': False, 'error': 'Error de conexión'}), 500
    
    cursor = conn.cursor()
    
    try:
        # Construir filtros
        filtro_fecha = "EXTRACT(YEAR FROM e.fecha_ingreso) = %s"
        params = [anio]
        
        if mes:
            filtro_fecha += " AND EXTRACT(MONTH FROM e.fecha_ingreso) = %s"
            params.append(mes)
        
        filtro_categoria = ""
        if categoria:
            filtro_categoria = " AND s.categoria LIKE %s"
            params.append(f'%{categoria}%')
        
        # 1. Estadísticas generales
        cursor.execute(f"""
            SELECT 
                COUNT(*) as total,
                SUM(CASE WHEN e.estado = 'Finalizado' THEN 1 ELSE 0 END) as finalizados,
                SUM(CASE WHEN e.estado = 'Pendiente' THEN 1 ELSE 0 END) as pendientes
            FROM equipos e
            LEFT JOIN solicitudes s ON e.solicitud_id = s.id
            WHERE {filtro_fecha} {filtro_categoria}
            AND e.eliminado = FALSE
        """, params.copy())
        
        stats = cursor.fetchone()
        print(f"DEBUG - Stats: {dict(stats) if stats else None}")
        
        # 2. Ingresos por mes
        cursor.execute(f"""
            SELECT 
                EXTRACT(YEAR FROM e.fecha_ingreso)::integer as anio,
                EXTRACT(MONTH FROM e.fecha_ingreso)::integer as mes,
                COUNT(*) as ingresados,
                SUM(CASE WHEN e.estado = 'Finalizado' THEN 1 ELSE 0 END) as finalizados,
                SUM(CASE WHEN e.estado = 'En curso' THEN 1 ELSE 0 END) as en_curso,
                SUM(CASE WHEN e.estado = 'Pendiente' THEN 1 ELSE 0 END) as pendientes,
                SUM(CASE WHEN e.estado NOT IN ('Finalizado', 'En curso', 'Pendiente') THEN 1 ELSE 0 END) as otros
            FROM equipos e
            LEFT JOIN solicitudes s ON e.solicitud_id = s.id
            WHERE {filtro_fecha} {filtro_categoria}
            AND e.eliminado = FALSE
            AND e.fecha_ingreso IS NOT NULL
            GROUP BY EXTRACT(YEAR FROM e.fecha_ingreso), EXTRACT(MONTH FROM e.fecha_ingreso)
            ORDER BY anio, mes
        """, params.copy())
        
        ingresos_por_mes = cursor.fetchall()
        print(f"DEBUG - Ingresos por mes ({len(ingresos_por_mes)} filas):")
        for i, row in enumerate(ingresos_por_mes[:3]):  # Solo primeras 3 filas
            print(f"  Fila {i}: {dict(row)}")
        
        # 3. Estados
        cursor.execute(f"""
            SELECT 
                e.estado,
                COUNT(*) as cantidad
            FROM equipos e
            LEFT JOIN solicitudes s ON e.solicitud_id = s.id
            WHERE {filtro_fecha} {filtro_categoria}
            AND e.eliminado = FALSE
            GROUP BY e.estado
            ORDER BY cantidad DESC
        """, params.copy())
        
        estados = cursor.fetchall()
        print(f"DEBUG - Estados ({len(estados)} filas)")
        
        # 4. Categorías
        cursor.execute(f"""
            SELECT 
                CASE 
                    WHEN s.categoria LIKE '%%R%%' THEN 'Reparación'
                    WHEN s.categoria LIKE '%%G%%' THEN 'Garantía'
                    WHEN s.categoria LIKE '%%BA%%' THEN 'Baja de Alquiler'
                    WHEN s.categoria LIKE '%%CA%%' THEN 'Cambio de Alquiler'
                    WHEN s.categoria LIKE '%%FC%%' THEN 'Cambio por Falla Crítica'
                    ELSE 'Otra'
                END as categoria,
                COUNT(*) as cantidad
            FROM equipos e
            LEFT JOIN solicitudes s ON e.solicitud_id = s.id
            WHERE {filtro_fecha} {filtro_categoria}
            AND e.eliminado = FALSE
            GROUP BY categoria
            ORDER BY cantidad DESC
        """, params.copy())
        
        categorias = cursor.fetchall()
        print(f"DEBUG - Categorías ({len(categorias)} filas)")
        
        cursor.close()
        conn.close()
        
        # Convertir los datos a formato JSON-serializable
        ingresos_list = []
        for i, row in enumerate(ingresos_por_mes):
            try:
                # Usar .get() en lugar de [] para evitar KeyError
                ingresos_list.append({
                    'anio': int(row.get('anio', anio)) if row.get('anio') else anio,
                    'mes': int(row.get('mes', 1)) if row.get('mes') else 1,
                    'ingresados': int(row.get('ingresados', 0)) if row.get('ingresados') else 0,
                    'finalizados': int(row.get('finalizados', 0)) if row.get('finalizados') else 0,
                    'en_curso': int(row.get('en_curso', 0)) if row.get('en_curso') else 0,
                    'pendientes': int(row.get('pendientes', 0)) if row.get('pendientes') else 0,
                    'otros': int(row.get('otros', 0)) if row.get('otros') else 0
                })
            except Exception as e:
                print(f"Error procesando fila {i} de ingresos_por_mes: {e}")
                print(f"Datos de la fila: {dict(row)}")
                raise
        
        estados_list = []
        for i, row in enumerate(estados):
            try:
                estados_list.append({
                    'estado': str(row.get('estado', 'Sin estado')) if row.get('estado') else 'Sin estado',
                    'cantidad': int(row.get('cantidad', 0)) if row.get('cantidad') else 0
                })
            except Exception as e:
                print(f"Error procesando fila {i} de estados: {e}")
                print(f"Datos de la fila: {dict(row)}")
                raise
        
        categorias_list = []
        for i, row in enumerate(categorias):
            try:
                categorias_list.append({
                    'categoria': str(row.get('categoria', 'Otra')) if row.get('categoria') else 'Otra',
                    'cantidad': int(row.get('cantidad', 0)) if row.get('cantidad') else 0
                })
            except Exception as e:
                print(f"Error procesando fila {i} de categorias: {e}")
                print(f"Datos de la fila: {dict(row)}")
                raise
        
        return jsonify({
            'success': True,
            'total': int(stats.get('total', 0)) if stats and stats.get('total') else 0,
            'total_ingresados': int(stats.get('total', 0)) if stats and stats.get('total') else 0,
            'total_finalizados': int(stats.get('finalizados', 0)) if stats and stats.get('finalizados') else 0,
            'total_pendientes': int(stats.get('pendientes', 0)) if stats and stats.get('pendientes') else 0,
            'ingresos_por_mes': ingresos_list,
            'estados': estados_list,
            'categorias': categorias_list
        })
    
    except Exception as e:
        print(f"Error al generar informe: {e}")
        import traceback
        traceback.print_exc()
        if cursor:
            cursor.close()
        if conn:
            conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/informe-mensual/excel')
@login_required
def api_informe_mensual_excel():
    """Exportar informe mensual a Excel"""
    try:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import send_file
        
        anio = request.args.get('anio', type=int)
        mes = request.args.get('mes', type=int)
        categoria = request.args.get('categoria', '')
        
        if not anio:
            return jsonify({'error': 'Año requerido'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener datos (reutilizar lógica del endpoint anterior)
        filtro_fecha = "EXTRACT(YEAR FROM e.fecha_ingreso) = %s"
        params = [anio]
        
        if mes:
            filtro_fecha += " AND EXTRACT(MONTH FROM e.fecha_ingreso) = %s"
            params.append(mes)
        
        filtro_categoria = ""
        if categoria:
            filtro_categoria = " AND s.categoria LIKE %s"
            params.append(f'%{categoria}%')
        
        # Obtener todos los equipos del período
        cursor.execute(f"""
            SELECT 
                e.ost,
                e.cliente,
                e.estado,
                TO_CHAR(e.fecha_ingreso, 'DD/MM/YYYY') as fecha_ingreso,
                e.tipo_equipo,
                e.marca,
                e.modelo,
                s.categoria,
                COALESCE(s.comercial_syemed, s.solicitante) as comercial
            FROM equipos e
            LEFT JOIN solicitudes s ON e.solicitud_id = s.id
            WHERE {filtro_fecha} {filtro_categoria}
            AND e.eliminado = FALSE
            ORDER BY e.fecha_ingreso DESC
        """, params)
        
        equipos = cursor.fetchall()
        
        cursor.close()
        conn.close()
        
        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Informe Mensual"
        
        # Estilos
        header_fill = PatternFill(start_color="1E4D7B", end_color="1E4D7B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = ['OST', 'Cliente', 'Estado', 'Fecha Ingreso', 'Tipo Equipo', 
                  'Marca', 'Modelo', 'Categoría', 'Comercial']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Datos
        for row_idx, equipo in enumerate(equipos, 2):
            ws.cell(row=row_idx, column=1, value=equipo['ost'])
            ws.cell(row=row_idx, column=2, value=equipo['cliente'])
            ws.cell(row=row_idx, column=3, value=equipo['estado'])
            ws.cell(row=row_idx, column=4, value=equipo['fecha_ingreso'])
            ws.cell(row=row_idx, column=5, value=equipo['tipo_equipo'])
            ws.cell(row=row_idx, column=6, value=equipo['marca'])
            ws.cell(row=row_idx, column=7, value=equipo['modelo'])
            ws.cell(row=row_idx, column=8, value=equipo['categoria'])
            ws.cell(row=row_idx, column=9, value=equipo['comercial'])
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 25
        
        # Guardar en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"informe_mensual_{anio}"
        if mes:
            filename += f"_{mes:02d}"
        filename += ".xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except ImportError:
        return jsonify({'error': 'openpyxl no está instalado. Ejecuta: pip install openpyxl'}), 500
    except Exception as e:
        print(f"Error al exportar Excel: {e}")
        return jsonify({'error': str(e)}), 500

# ============================================
# CONTEXT PROCESSOR PARA TEMPLATES
# ============================================

@app.context_processor
def inject_user():
    """Inyecta información del usuario actual en todos los templates"""
    return dict(current_user=current_user)

if __name__ == '__main__':
    app.run(debug=True)